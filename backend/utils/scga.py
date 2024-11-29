# ****************************************************
# run pip install -r requirements.txt for dependencies 
# ****************************************************
import openpyxl
import traceback
import datetime

NULL = ''
# scga class
INCOMPLETE_TESTS = 'INCOMPLETE_TESTS'
REQUIREMENTS_CODE_MISMATCH = 'REQUIREMENTS_CODE_MISMATCH'
DEACTIVATED_CODE = 'DEACTIVATED_CODE'
DEFENSIVE_CODE = 'DEFENSIVE_CODE'
TEST_ENVIRONMENT_LIMITATIONS = 'TEST_ENVIRONMENT_LIMITATIONS'
PREVIOUSLY_ANALYZED_SOFTWARE = 'PREVIOUSLY_ANALYZED_SOFTWARE'
OTHER = 'OTHER'
CHOICES_CLASS = (
    (NULL, None),
    (INCOMPLETE_TESTS, "Incomplete Tests"),
    (REQUIREMENTS_CODE_MISMATCH, "Requirements-Code Mismatch"),
    (DEACTIVATED_CODE, "Deactivated Code"),
    (DEFENSIVE_CODE, "Defensive Code"),
    (TEST_ENVIRONMENT_LIMITATIONS, "Test Environment Limitations"),
    (PREVIOUSLY_ANALYZED_SOFTWARE, "Previously Analyzed Software"),
    (OTHER, "Other"),
)


def set_class(value):
    choice_map = {v: k for k, v in CHOICES_CLASS}
    class_choice = None
    if value in choice_map:
        class_choice = choice_map[value]
    else:
        class_choice = ''
    return class_choice


class SCGA:
    TEST_PLAN = 'plan'
    TEST_EXCEPTIONS = 'exceptions'
    LEVEL = 'Level'

    def __init__(self, scga_f, info = None):
        if not(scga_f.endswith('xlsm') and '~' not in str(scga_f) and 'SCGA' in str(scga_f)):
            print("Inpropre input file type: must be a .xlsm file with 'SCGA' included in filename")
            raise ValueError
        self.scga_wb = openpyxl.load_workbook(scga_f, read_only=True)
        self.scga_dict = {
            'project': '',
            'function': '',
            'current': '',
            'file_name': str(scga_f).split("\\")[-1],
            'baseline': str(scga_f).split("\\")[-1].split('_SCGA')[0],
            'levels': []
        }
        str(scga_f).split("/")[-1]
        self.scga_function_list = {
            'baseline': str(scga_f).split('_SCGA')[0],
            'levelAFunctions': [],
            'levelBFunctions': [],
            'levelCFunctions': []
        }
        if info:
            self.scga_dict['project'] = info['project']
            self.scga_dict['function'] = info['function']
            if info['current']: self.scga_dict['current'] = 'Y'
            else: self.scga_dict['current'] = 'N'


    def get_raw_data(sheet, min_row, max_col):
        raw_data = []
        for row in sheet.iter_rows(min_row=min_row, max_col=max_col, values_only=True):
            raw_data.append(row)
        return raw_data


    def read_scga(self):
        read_plan = False
        read_exception = False
        # scga_sheets = scga_workbook.worksheets
        for cur_sheet in self.scga_wb:
            if SCGA.LEVEL in cur_sheet.title: # locate SCGA result
                # only create new one if read_plan and read_exception is false(not yet read)
                if not (read_plan or read_exception):
                    Level = {
                        'level': str(cur_sheet.title).split(' ')[1],
                        'test_plan': {},
                        'test_exception': {},
                    }
                # locate to file end
                # scga_log_f.seek(0, 2)
                
                if (SCGA.TEST_PLAN in str(cur_sheet.title).lower()) and ('is not affect' not in cur_sheet['A7'].value): # locate valid test plan
                    test_plan = {
                        'sheet_name': cur_sheet.title,
                        'modules': {},
                        'lv_total_coverage': {}
                    }
                    
                    test_plan['modules'], test_plan['lv_total_coverage'], function_list = self.read_test_plan(cur_sheet, Level['level'])
                    Level['test_plan'] = test_plan
                    
                    if Level['level'] == 'A':
                        # SCGA['level_A']['test_plan'] = test_plan
                        self.scga_function_list['levelAFunctions'] = function_list
                    elif Level['level'] == 'B':
                        # SCGA['level_B']['test_plan'] = test_plan
                        self.scga_function_list['levelBFunctions'] = function_list
                    elif Level['level'] == 'C':
                        # SCGA['level_C']['test_plan'] = test_plan
                        self.scga_function_list['levelCFunctions'] = function_list
                    read_plan = True
                if (SCGA.TEST_EXCEPTIONS in str(cur_sheet.title).lower()) and ('is not affect' not in cur_sheet['A6'].value): # locate valid test exceptions
                    test_exception = {
                        'sheet_name': cur_sheet.title,
                        # 'level': None,
                        'modules': [],
                    }
                    test_exception['modules'] = self.read_test_exceptions(cur_sheet)
                    Level['test_exception'] = test_exception
                    read_exception = True
                # else:
                    # output_log(
                    #     f'* SCGA information not found in Level {Level["level"]}')
                # append if plan & exceptions are both read
                # can only put after read of plan and exception to seeks the first moment that bith plan & exceptions are both read
                if (read_plan and read_exception): # append into scga
                    self.scga_dict['levels'].append(Level)
                    read_plan = False
                    read_exception = False
                    

    def read_info_behind_coverages(self, row_data, level):
        covered = {}
        total = {}
        if level == 'A':
        # covered module strcuture data
            covered['branches'] = int(row_data[11])
            covered['pairs'] = int(row_data[12])
            covered['statement'] = int(row_data[13])

            # total module strcuture data
            total['branches'] = int(row_data[14])
            total['pairs'] = int(row_data[15])
            total['statement'] = int(row_data[16])
        elif level == 'B':
            # covered module strcuture data
            covered['branches'] = int(row_data[11])
            covered['statement'] = int(row_data[12])

            # total module strcuture data
            total['branches'] = int(row_data[13])
            total['statement'] = int(row_data[14])
        elif level == 'C':
            # total module strcuture data
            total['statement'] = int(row_data[11])

            # covered module strcuture data
            covered['statement'] = int(row_data[10])
        
        return covered, total,


    def read_test_plan(self, tpsheet, level):
        """Read test plan inforamtions of scga sheet

        Args:
            sheet (raw data): test plan sheet of a scga excel
            rows (int): row numbers of test plan sheet

        Raises:
            ValueError: None

        Returns:
            dict list -- module: dict list of all files involved in scga  
            dict -- level_total_coverage: coverage condition of scga 
            list -- functions_name_list: functions involed in scga
        """
        tp_start_row = 7
        tp_max_col = 0
        if level == 'A': tp_max_col = 21
        elif level == 'B': tp_max_col = 19
        elif level == 'C':  tp_max_col = 17
        modules = []
        modules_name_list = []
        functions_name_list = []
        total_function = 0
        level_total_coverage = {
            'percent_coverage_MCDC': 0,
            'percent_coverage_Analysis': 0,
            'total_coverage': 0,
        }
        module = {
            'module_name': None,
            'process': None,
            'functions': []
        }
        try:
            for tp_row_data in tpsheet.iter_rows(min_row=tp_start_row, max_col=tp_max_col, values_only=True):
                if tp_row_data[1] is None:
                    if 'level' in str(tp_row_data[0]).lower():
                        level_total_coverage['percent_coverage_MCDC'] = tp_row_data[7]
                        level_total_coverage['percent_coverage_Analysis'] = tp_row_data[8]
                        level_total_coverage['total_coverage'] = tp_row_data[9]
                        continue
                    else:
                        continue
                function = {
                    'function_name': None,
                    'analyst': None,
                    'site': None,
                    'start_date': None,
                    'coverage': {},
                    'covered': {},
                    'total': {},
                    # 'moduleStrucData': {},
                    'oversight': None,
                    'defect_classification': {}
                }
                coverage = {
                    'percent_coverage_MCDC': 0,
                    'percent_coverage_Analysis': 0,
                    'total_coverage': 0,
                }
                defect_classification  = {
                    'tech': '',
                    'non_tech': '',
                    'process': '',
                }
                
                # see if tp_row_data a new module line
                if module['module_name'] != tp_row_data[1]:

                    if module['module_name'] is not None and module['module_name'] not in modules_name_list:
                        modules_name_list.append(module['module_name'])
                        modules.append(module)
                    module = {
                        'module_name': None,
                        'functions': []
                    }
                    module['process'] = tp_row_data[0]
                    module['module_name'] = tp_row_data[1]
                
                # defect _classification

                defect_classification['tech'] = tp_row_data[-3]
                defect_classification['non_tech'] = tp_row_data[-2]
                defect_classification['process'] = tp_row_data[-1]
                function['defect_classification'] = defect_classification

                function['covered'], function['total'] = self.read_info_behind_coverages(tp_row_data, level)
                
                # coverage
                coverage['percent_coverage_MCDC'] = tp_row_data[7]
                coverage['percent_coverage_Analysis'] = tp_row_data[8]
                coverage['total_coverage'] = tp_row_data[9]
                # function
                function['function_name'] = tp_row_data[2]
                function['analyst'] = tp_row_data[4]
                function['site'] = tp_row_data[5]
                # strfy datetime
                if isinstance(tp_row_data[6], datetime.datetime):
                    function['start_date'] = tp_row_data[6].strftime("%Y-%m-%d")
                else:
                    function['start_date'] = tp_row_data[6]
                function['coverage'] = coverage

                # see if tp_row_data a new module line
                if tp_row_data[1] not in modules_name_list or len(modules) == 0:
                    module = {
                        'module_name': None,
                        'functions': []
                    }
                    module['process'] = tp_row_data[0]
                    module['module_name'] = tp_row_data[1]
                    module['functions'].append(function)
                    functions_name_list.append(function['function_name'])
                    total_function = total_function + 1
                    modules.append(module)
                    modules_name_list.append(module['module_name'])
                # in case module already exist, add function to corresponding module
                else:
                    for index, module_name in enumerate(modules_name_list):
                        if module_name == str(module['module_name']):
                            (modules[index])['functions'].append(function)
                            functions_name_list.append(function['function_name'])
                            total_function = total_function + 1
        except Exception as e:
            import pdb; pdb.set_trace()
            print(traceback.print_exc())
        # write into log file
        # output_log(
        #     f"Total functions of {tpsheet.title} is: {total_function}")
        # output_log(f"total functions shows as below ({len(functions_name_list)}):")
        # output_log(functions_name_list)
        # output_log(f"total module shows as below ({len(modules_name_list)}):")
        # output_log(modules_name_list)
        return modules, level_total_coverage, functions_name_list


    def read_test_exceptions(self, tesheet):
        """Read test exception informations of scga

        Args:
            tesheet (xlwing test exception sheet): xlwing test exception sheet
            rows (int): row number of test exception sheet

        Raises:
            ValueError: None

        Returns:
            dict list -- uncovered_modules: dataset of all uncovered file involved in scga
        """
        te_start_row = 6
        te_max_col = 13
        uncovered_modules = []
        uncovered_modules_name_list = []
        uncovered_functions_name_list = []
        uncovered_module = {
            'module_name': None,
            'functions': []
        }
        uncovered_function = {
            'function_name': None,
            'note': None,
            'uncoverage': []
        }
        total_uncoverage = 0
        for te_row_data in tesheet.iter_rows(min_row=te_start_row, max_col=te_max_col, values_only=True):
            if te_row_data[1] is None:
                continue
            uncoverage = {
                'uncovered_sw_line': [],
                'uncovered_instrument_sw_line': [],
                'requirement_id': '',
                'analyst': None,
                '_class': None,
                '_class_str': None,
                'analysis_summary': None,
                'correction_summary': None,
                'issue': None,
                'PAR_SCR': None,
                'comment': None,
            }
            uncovered_function['note'] = te_row_data[0]
            # uncovered software line
            uncoverage['uncovered_sw_line'] = te_row_data[3]
            # uncovered software code
            uncoverage['uncovered_instrument_sw_line'] = te_row_data[4]
            # requirements
            if te_row_data[5] is not None:
                # uncoverage['requirement_id'] = str(info[5]).split('\n')
                uncoverage['requirement_id'] = te_row_data[5]
            uncoverage['analyst'] = te_row_data[6]
            uncoverage['_class_str'] = str(te_row_data[7]).replace('\\', '')
            uncoverage['_class'] = set_class(te_row_data[7])
            uncoverage['correction_summary'] = te_row_data[9]
            uncoverage['issue'] = te_row_data[10]
            uncoverage['PAR_SCR'] = te_row_data[11]
            uncoverage['comment'] = te_row_data[12]
            # applicable['PAR_SCR'] = info[11]
            # applicable['comment'] = info[12]
            # uncoverage['applicable'] = applicable
            # see if info has new function
            if te_row_data[2] not in uncovered_functions_name_list or len(uncovered_functions_name_list) == 0:
                uncovered_function = {
                    'function_name': None,
                    'note': None,
                    'uncoverages': [],
                    'uncoverage_count': 0
                }
                uncovered_function['note'] = te_row_data[0]
                uncovered_function['function_name'] = te_row_data[2]
                uncovered_function['uncoverages'].append(uncoverage)
                uncovered_function['uncoverage_count'] = uncovered_function['uncoverage_count'] + 1
                total_uncoverage = total_uncoverage + 1
                uncovered_functions_name_list.append(
                    uncovered_function['function_name'])
                # see if info a new module line
                if te_row_data[1] not in uncovered_modules_name_list or len(uncovered_modules) == 0:
                    uncovered_module = {
                        'module_name': None,
                        'functions': []
                    }
                    uncovered_module['process'] = te_row_data[0]
                    uncovered_module['module_name'] = te_row_data[1]
                    uncovered_module['functions'].append(uncovered_function)
                    uncovered_modules.append(uncovered_module)
                    uncovered_modules_name_list.append(
                        uncovered_module['module_name'])
                # in case module already exist, add function to corresponding module
                else:
                    for index, module_name in enumerate(uncovered_modules_name_list):
                        if module_name == str(uncovered_module['module_name']):
                            (uncovered_modules[index])[
                                'functions'].append(uncovered_function)
            else:  # in case function already exist, add uncoverage information to corresponding function in corresponding module
                for i, module_name in enumerate(uncovered_modules_name_list):
                    if module_name == te_row_data[1]:
                        for j, uncovered_function in enumerate((uncovered_modules[i])['functions']):
                            if te_row_data[2] == uncovered_function['function_name']:
                                (uncovered_modules[i])[
                                    'functions'][j]['uncoverages'].append(uncoverage)
                                (uncovered_modules[i])['functions'][j]['uncoverage_count'] = (
                                    uncovered_modules[i])['functions'][j]['uncoverage_count'] + 1
                                total_uncoverage = total_uncoverage + 1
        # write into log file
        # output_log(
        #     f"Total uncovered situation of {tesheet.name} is: {total_uncoverage}")
        # output_log(
        #     f"total functions shows as below ({len(uncovered_functions_name_list)}):")
        # output_log(uncovered_functions_name_list)
        # output_log(
        #     f"total module shows as below ({len(uncovered_modules_name_list)}):")
        # output_log(uncovered_modules_name_list)
        return uncovered_modules


def test():
    scga_f = input("Please enter the file path: ")
    # scga = SCGA(scga_f)
    scga = SCGA(scga_f)
    scga.read_scga()
    print(scga.scga_function_list)



if __name__ == "__main__":
    test()